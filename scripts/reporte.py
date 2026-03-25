"""
Genera reporte diario y acumulado desde MongoDB, Excel, PNG y envío por correo.
Pensado para ejecución local y en GitHub Actions (sin persistencia local entre corridas).
"""

from __future__ import annotations

import logging
import os
import smtplib
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from typing import Any, Iterable

if "MPLCONFIGDIR" not in os.environ:
    _mpl_dir = os.path.join(tempfile.gettempdir(), "matplotlib-cache")
    os.makedirs(_mpl_dir, exist_ok=True)
    os.environ["MPLCONFIGDIR"] = _mpl_dir

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.offsetbox import AnnotationBbox, OffsetImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pymongo import MongoClient
from pymongo.collection import Collection

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

LOG = logging.getLogger("reporte")


@dataclass(frozen=True)
class Config:
    mongo_uri: str
    db_name: str
    lead_collection: str
    loans_collection: str
    users_collection: str
    ecosystem_id: str
    email_remitente: str
    email_password: str
    email_destinos: list[str]
    fecha_lanzamiento: date


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise ValueError(f"Falta la variable de entorno obligatoria: {name}")
    return value


def _parse_fecha_lanzamiento(raw: str) -> date:
    raw = raw.strip()
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(
            "FECHA_LANZAMIENTO debe ser YYYY-MM-DD (ej. 2026-03-02)"
        ) from exc


def load_config() -> Config:
    # Si el secreto no existe en GitHub Actions, la variable puede venir vacía.
    fecha_raw = (os.environ.get("FECHA_LANZAMIENTO") or "").strip() or "2026-03-02"
    destinos_raw = _require_env("EMAIL_DESTINO")
    destinos = [d.strip() for d in destinos_raw.split(",") if d.strip()]
    if not destinos:
        raise ValueError("EMAIL_DESTINO debe contener al menos un correo (separados por comas).")

    return Config(
        mongo_uri=_require_env("MONGO_URI"),
        db_name=_require_env("DB_NAME"),
        lead_collection=_require_env("LEAD_COLLECTION"),
        loans_collection=_require_env("LOANS_COLLECTION"),
        users_collection=_require_env("USERS_COLLECTION"),
        ecosystem_id=_require_env("ECOSYSTEM_ID"),
        email_remitente=_require_env("EMAIL_REMITENTE"),
        email_password=_require_env("EMAIL_PASSWORD"),
        email_destinos=destinos,
        fecha_lanzamiento=_parse_fecha_lanzamiento(fecha_raw),
    )


def script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def logo_path() -> str:
    return os.path.join(script_dir(), "logos", "guatecredito.png")


def load_logo_for_matplotlib():
    path = logo_path()
    if not os.path.isfile(path):
        LOG.warning("No se encontró %s; los gráficos irán sin logo.", path)
        return None
    return plt.imread(path)


def guatemala_tz() -> timezone:
    return timezone(timedelta(hours=-6))


def date_to_day_bounds_utc_iso(d: date) -> tuple[str, str]:
    """Inicio y fin del día calendario `d` en zona Guatemala, expresados en ISO UTC Z."""
    gt = guatemala_tz()
    inicio_gt = datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=gt)
    fin_gt = inicio_gt + timedelta(days=1) - timedelta(seconds=1)
    inicio_utc = inicio_gt.astimezone(timezone.utc)
    fin_utc = fin_gt.astimezone(timezone.utc)
    return (
        inicio_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        fin_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
    )


def add_logo(ax, logo, x: float, y: float, zoom: float = 0.22) -> None:
    if logo is None:
        return
    imagebox = OffsetImage(logo, zoom=zoom)
    ab = AnnotationBbox(imagebox, (x, y), frameon=False, xycoords="axes fraction")
    ax.add_artist(ab)


def _rechazo_politicas_del_dia(lead_col: Collection, ecosystem_id: str, inicio_iso: str, fin_iso: str) -> int:
    pipeline = [
        {
            "$match": {
                "ecosystem": ecosystem_id,
                "created_at": {"$gte": inicio_iso, "$lte": fin_iso},
            }
        },
        {"$unwind": "$additional_data.variables_data"},
        {
            "$match": {
                "additional_data.variables_data.key": {"$in": ["motivos_negacion", "motivoNegacion"]}
            }
        },
        {"$group": {"_id": "$additional_data.variables_data.value", "cantidad": {"$sum": 1}}},
    ]
    total = 0
    for r in lead_col.aggregate(pipeline):
        motivo = (r["_id"] or "").lower()
        cantidad = r["cantidad"]
        if (
            "no cumple" in motivo
            or "identificación" in motivo
            or "nis" in motivo
            or "lista negra" in motivo
        ):
            total += cantidad
    return total


def _loans_summary(loans_col: Collection, ecosystem_id: str, inicio_iso: str, fin_iso: str) -> dict[str, Any]:
    pipeline = [
        {
            "$match": {
                "financial_entity_id": ecosystem_id,
                "assignment_date": {"$gte": inicio_iso, "$lte": fin_iso},
            }
        },
        {
            "$group": {
                "_id": None,
                "creditos_nuevos": {"$sum": 1},
                "desembolsos": {"$sum": {"$cond": [{"$eq": ["$disbursed", True]}, 1, 0]}},
                "no_desembolsos": {"$sum": {"$cond": [{"$eq": ["$disbursed", False]}, 1, 0]}},
                "monto_desembolsado": {"$sum": {"$cond": [{"$eq": ["$disbursed", True]}, "$amount", 0]}},
            }
        },
    ]
    rows = list(loans_col.aggregate(pipeline))
    if rows:
        return rows[0]
    return {
        "creditos_nuevos": 0,
        "desembolsos": 0,
        "no_desembolsos": 0,
        "monto_desembolsado": 0,
    }


def compute_metrics_for_day(
    lead_col: Collection,
    users_col: Collection,
    loans_col: Collection,
    ecosystem_id: str,
    dia: date,
) -> dict[str, Any]:
    inicio_iso, fin_iso = date_to_day_bounds_utc_iso(dia)
    fecha_str = dia.strftime("%Y-%m-%d")

    prospectos = lead_col.count_documents(
        {"ecosystem": ecosystem_id, "created_at": {"$gte": inicio_iso, "$lte": fin_iso}}
    )
    clientes = users_col.count_documents(
        {"ecosystem": ecosystem_id, "created_at": {"$gte": inicio_iso, "$lte": fin_iso}}
    )
    rechazo_politicas = _rechazo_politicas_del_dia(lead_col, ecosystem_id, inicio_iso, fin_iso)
    data_loans = _loans_summary(loans_col, ecosystem_id, inicio_iso, fin_iso)

    creditos = data_loans["creditos_nuevos"]
    desembolsos = data_loans["desembolsos"]
    no_desembolsos = data_loans["no_desembolsos"]
    monto = data_loans["monto_desembolsado"] / 100
    ticket = round(monto / desembolsos, 2) if desembolsos else 0
    solicitudes = prospectos + clientes
    rechazos_bd = prospectos - rechazo_politicas

    return {
        "Fecha": fecha_str,
        "Solicitudes": solicitudes,
        "Rechazos - BD": rechazos_bd,
        "Clientes nuevos": clientes,
        "Rechazos - Políticas": rechazo_politicas,
        "Créditos nuevos": creditos,
        "Desembolsos": desembolsos,
        "No desembolsos": no_desembolsos,
        "Monto desembolsado": monto,
        "Ticket promedio": ticket,
    }


def build_dataframe_historico_mongo(
    lead_col: Collection,
    users_col: Collection,
    loans_col: Collection,
    ecosystem_id: str,
    fecha_inicio: date,
    fecha_fin: date,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    d = fecha_inicio
    while d <= fecha_fin:
        rows.append(compute_metrics_for_day(lead_col, users_col, loans_col, ecosystem_id, d))
        d += timedelta(days=1)
    return pd.DataFrame(rows)


def compute_acumulado_mongo(
    lead_col: Collection,
    users_col: Collection,
    loans_col: Collection,
    ecosystem_id: str,
    inicio_acum_iso: str,
    fin_acum_iso: str,
) -> dict[str, Any]:
    prospectos = lead_col.count_documents(
        {"ecosystem": ecosystem_id, "created_at": {"$gte": inicio_acum_iso, "$lte": fin_acum_iso}}
    )
    clientes = users_col.count_documents(
        {"ecosystem": ecosystem_id, "created_at": {"$gte": inicio_acum_iso, "$lte": fin_acum_iso}}
    )
    rechazo_politicas = _rechazo_politicas_del_dia(lead_col, ecosystem_id, inicio_acum_iso, fin_acum_iso)
    data_loans = _loans_summary(loans_col, ecosystem_id, inicio_acum_iso, fin_acum_iso)

    creditos = data_loans["creditos_nuevos"]
    desembolsos = data_loans["desembolsos"]
    no_desembolsos = data_loans["no_desembolsos"]
    monto = data_loans["monto_desembolsado"] / 100
    ticket = round(monto / desembolsos, 2) if desembolsos else 0
    solicitudes = prospectos + clientes
    rechazos_bd = prospectos - rechazo_politicas

    return {
        "Solicitudes": solicitudes,
        "Rechazos - BD": rechazos_bd,
        "Clientes nuevos": clientes,
        "Rechazos - Políticas": rechazo_politicas,
        "Créditos nuevos": creditos,
        "Desembolsos": desembolsos,
        "No desembolsos": no_desembolsos,
        "Monto desembolsado": monto,
        "Ticket promedio": ticket,
    }


def format_excel_workbook(path: str, fecha_ejecucion: str) -> None:
    wb = load_workbook(path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = max_length + 2

    ws["C1"] = "Reporte acumulado Guatecrédito"
    ws["C1"].font = Font(bold=True)
    ws["C2"] = f"Fecha de ejecución: {fecha_ejecucion}"

    lp = logo_path()
    if os.path.isfile(lp):
        try:
            logo_excel = XLImage(lp)
            logo_excel.width = 120
            logo_excel.height = 60
            ws.add_image(logo_excel, "A1")
        except Exception:
            LOG.exception("No se pudo insertar el logo en el Excel.")

    wb.save(path)


def render_report_image(
    title: str,
    subtitle_lines: Iterable[str],
    labels: list[str],
    values: list[Any],
    out_path: str,
    logo,
) -> None:
    fig, ax = plt.subplots(figsize=(6, 8))
    ax.axis("off")
    ax.text(0.5, 0.96, title, ha="center", fontsize=18, weight="bold")
    y_sub = 0.92
    for line in subtitle_lines:
        ax.text(0.5, y_sub, line, ha="center", fontsize=11)
        y_sub -= 0.04
    add_logo(ax, logo, 0.08, 0.95)
    add_logo(ax, logo, 0.92, 0.95)
    y = 0.72
    for label, value in zip(labels, values):
        ax.text(0.10, y, label, fontsize=11, weight="bold")
        ax.text(0.80, y, str(value), ha="right", fontsize=11)
        y -= 0.06
    plt.subplots_adjust(top=0.92)
    fig.savefig(out_path, dpi=300)
    plt.close(fig)


def send_email_gmail_ssl(
    remitente: str,
    password: str,
    destinos: list[str],
    asunto: str,
    cuerpo: str,
    adjuntos: list[tuple[str, str, str, bytes]],
) -> None:
    """
    adjuntos: lista de (maintype, subtype, filename, data)
    """
    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = remitente
    msg["To"] = ", ".join(destinos)
    msg.set_content(cuerpo)
    for maintype, subtype, filename, data in adjuntos:
        msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=filename)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=120) as smtp:
        smtp.login(remitente, password)
        smtp.send_message(msg)


def main() -> int:
    setup_logging()
    if load_dotenv is not None:
        load_dotenv()

    try:
        cfg = load_config()
    except ValueError as e:
        LOG.error("%s", e)
        return 1

    gt = guatemala_tz()
    hoy = datetime.now(gt)
    fecha_consulta = hoy.date()
    fecha_consulta_str = fecha_consulta.strftime("%Y-%m-%d")
    hora_actual = hoy.strftime("%H%M%S")
    fecha_ejecucion = hoy.strftime("%Y-%m-%d %H:%M:%S")

    if fecha_consulta < cfg.fecha_lanzamiento:
        LOG.error(
            "La fecha actual en Guatemala (%s) es anterior a FECHA_LANZAMIENTO (%s).",
            fecha_consulta_str,
            cfg.fecha_lanzamiento,
        )
        return 1

    base_dir = os.path.join(script_dir(), "reportes")
    excel_dir = os.path.join(base_dir, "excel_acumulado")
    carpeta_dia = os.path.join(base_dir, fecha_consulta_str)
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(carpeta_dia, exist_ok=True)

    archivo_excel = os.path.join(carpeta_dia, "reporte_general.xlsx")
    imagen_diaria = os.path.join(carpeta_dia, "reporte_diario.png")
    imagen_acumulada = os.path.join(carpeta_dia, "reporte_acumulado.png")

    inicio_dia_iso, fin_dia_iso = date_to_day_bounds_utc_iso(fecha_consulta)
    inicio_acum_gt = datetime(
        cfg.fecha_lanzamiento.year,
        cfg.fecha_lanzamiento.month,
        cfg.fecha_lanzamiento.day,
        0,
        0,
        0,
        tzinfo=gt,
    )
    inicio_acum_iso = inicio_acum_gt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    LOG.info("Rango día actual (UTC): %s .. %s", inicio_dia_iso, fin_dia_iso)
    LOG.info("Acumulado desde %s hasta fin del día %s (UTC inicio acum: %s)", cfg.fecha_lanzamiento, fecha_consulta_str, inicio_acum_iso)

    logo = load_logo_for_matplotlib()

    client: MongoClient | None = None
    try:
        client = MongoClient(cfg.mongo_uri, serverSelectionTimeoutMS=30_000)
        client.admin.command("ping")
        db = client[cfg.db_name]
        lead_col = db[cfg.lead_collection]
        loans_col = db[cfg.loans_collection]
        users_col = db[cfg.users_collection]
    except Exception:
        LOG.exception("Error conectando a MongoDB.")
        return 1

    try:
        df_final = build_dataframe_historico_mongo(
            lead_col,
            users_col,
            loans_col,
            cfg.ecosystem_id,
            cfg.fecha_lanzamiento,
            fecha_consulta,
        )
        df_final.to_excel(archivo_excel, index=False, startrow=4)
        format_excel_workbook(archivo_excel, fecha_ejecucion)

        fila_hoy = df_final[df_final["Fecha"] == fecha_consulta_str]
        if fila_hoy.empty:
            LOG.error("No hay fila para el día actual en el DataFrame; revisar fechas.")
            return 1
        row = fila_hoy.iloc[0]

        labels = [
            "Solicitudes",
            "Rechazos - BD",
            "Clientes nuevos",
            "Rechazos - Políticas",
            "Créditos nuevos",
            "Desembolsos",
            "No desembolsos",
            "Monto desembolsado",
            "Ticket promedio",
        ]
        values_dia = [
            int(row["Solicitudes"]),
            int(row["Rechazos - BD"]),
            int(row["Clientes nuevos"]),
            int(row["Rechazos - Políticas"]),
            int(row["Créditos nuevos"]),
            int(row["Desembolsos"]),
            int(row["No desembolsos"]),
            f"Q{float(row['Monto desembolsado']):,.2f}",
            f"Q{float(row['Ticket promedio']):,.2f}",
        ]

        render_report_image(
            "REPORTE DIARIO",
            [
                f"Fecha: {fecha_consulta_str}",
                f"Fecha de ejecución: {fecha_ejecucion}",
            ],
            labels,
            values_dia,
            imagen_diaria,
            logo,
        )

        acum = compute_acumulado_mongo(
            lead_col,
            users_col,
            loans_col,
            cfg.ecosystem_id,
            inicio_acum_iso,
            fin_dia_iso,
        )
        values_acum = [
            acum["Solicitudes"],
            acum["Rechazos - BD"],
            acum["Clientes nuevos"],
            acum["Rechazos - Políticas"],
            acum["Créditos nuevos"],
            acum["Desembolsos"],
            acum["No desembolsos"],
            f"Q{acum['Monto desembolsado']:,.2f}",
            f"Q{acum['Ticket promedio']:,.2f}",
        ]

        render_report_image(
            "ACUMULADO",
            [
                f"Desde: {cfg.fecha_lanzamiento} hasta: {fecha_consulta_str}",
                f"Fecha de ejecución: {fecha_ejecucion}",
            ],
            labels,
            values_acum,
            imagen_acumulada,
            logo,
        )
    except Exception:
        LOG.exception("Error generando reportes o Excel.")
        return 1
    finally:
        if client is not None:
            client.close()

    correo_ok = False
    try:
        with open(imagen_diaria, "rb") as f:
            png_diario = f.read()
        with open(imagen_acumulada, "rb") as f:
            png_acum = f.read()
        with open(archivo_excel, "rb") as f:
            xlsx_data = f.read()

        send_email_gmail_ssl(
            cfg.email_remitente,
            cfg.email_password,
            cfg.email_destinos,
            f"Reporte Guatecrédito {fecha_consulta_str}",
            (
                "Adjunto reporte diario y acumulado generado.\n\n"
                f"Fecha de consulta: {fecha_consulta_str}\n"
                f"Hora (Guatemala): {hora_actual}"
            ),
            [
                ("image", "png", "reporte_diario.png", png_diario),
                ("image", "png", "reporte_acumulado.png", png_acum),
                (
                    "application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "reporte_general.xlsx",
                    xlsx_data,
                ),
            ],
        )
        correo_ok = True
        LOG.info("Correo enviado correctamente a: %s", ", ".join(cfg.email_destinos))
    except Exception:
        LOG.exception("Error enviando correo.")

    if not correo_ok:
        return 1

    LOG.info("Excel: %s", archivo_excel)
    LOG.info("PNG diario: %s", imagen_diaria)
    LOG.info("PNG acumulado: %s", imagen_acumulada)
    return 0


if __name__ == "__main__":
    sys.exit(main())
